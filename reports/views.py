from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.db.models import Q, Count, Max
from datetime import datetime
import openpyxl
from django.template.loader import get_template
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from xhtml2pdf import pisa

from transactions.models import  SupportingDocument, SupportingDocumentFile, Comments

from main_app.models import Project
def project_report_view(request, project_id):
    """Enhanced view for displaying the report with filtering capabilities"""
    project = get_object_or_404(Project, project_id=project_id)

    # Get all supporting documents for this project
    supporting_docs_queryset = SupportingDocument.objects.filter(project=project)

    # Get filter parameters from request
    fiscal_year = request.GET.get('fiscal_year', '')
    fiscal_period = request.GET.get('fiscal_period', '')
    min_amount = request.GET.get('min_amount', '')
    max_amount = request.GET.get('max_amount', '')
    reference_number = request.GET.get('reference_number', '')
    support_status = request.GET.get('support_status', '')

    # Apply filters
    filtered_docs = supporting_docs_queryset

    if fiscal_year:
        filtered_docs = filtered_docs.filter(fiscal_year=fiscal_year)

    if fiscal_period:
        filtered_docs = filtered_docs.filter(fiscal_period=fiscal_period)

    if min_amount:
        try:
            min_amount_decimal = float(min_amount)
            filtered_docs = filtered_docs.filter(transaction_value__gte=min_amount_decimal)
        except ValueError:
            pass  # Invalid input, ignore filter

    if max_amount:
        try:
            max_amount_decimal = float(max_amount)
            filtered_docs = filtered_docs.filter(transaction_value__lte=max_amount_decimal)
        except ValueError:
            pass  # Invalid input, ignore filter

    if reference_number:
        filtered_docs = filtered_docs.filter(iddoc__icontains=reference_number.strip())

    if support_status:
        if support_status == 'supported':
            filtered_docs = filtered_docs.filter(supported=True)
        elif support_status == 'unsupported':
            filtered_docs = filtered_docs.filter(supported=False)

    # Order by fiscal year, period, then batch/entry
    filtered_docs = filtered_docs.select_related().prefetch_related('documents').order_by(
        '-fiscal_year', '-fiscal_period', 'batchnbr', 'entrynbr'
    )

    # Add latest comment to each document
    for doc in filtered_docs:
        latest_comment = Comments.objects.filter(
            project=project,
            batchnbr=doc.batchnbr,
            entrynbr=doc.entrynbr
        ).select_related('user').order_by('-timestamp').first()
        doc.latest_comment = latest_comment

    # Get filter options for dropdowns
    available_years = supporting_docs_queryset.values_list('fiscal_year', flat=True).distinct().order_by('-fiscal_year')
    available_periods = supporting_docs_queryset.values_list('fiscal_period', flat=True).distinct().order_by(
        'fiscal_period')

    # Get comments based on current filters
    filtered_comments = Comments.objects.filter(project=project)
    if filtered_docs.exists():
        batch_entry_combinations = [(doc.batchnbr, doc.entrynbr) for doc in filtered_docs]
        comment_filters = Q()
        for batchnbr, entrynbr in batch_entry_combinations:
            comment_filters |= Q(batchnbr=batchnbr, entrynbr=entrynbr)
        filtered_comments = filtered_comments.filter(comment_filters)

    filtered_comments = filtered_comments.select_related('user').order_by('-timestamp')

    # Calculate support percentage based on filtered data
    total_filtered = filtered_docs.count()
    supported_filtered = filtered_docs.filter(supported=True).count()
    support_percentage = (supported_filtered / total_filtered * 100) if total_filtered > 0 else 0

    # Prepare context for template
    context = {
        'project': project,
        'filtered_documents': filtered_docs,
        'total_documents': supporting_docs_queryset.count(),
        'filtered_comments': filtered_comments,
        'available_years': available_years,
        'available_periods': available_periods,
        'support_percentage': support_percentage,
        'report_date': datetime.now(),

        # Selected filter values (for maintaining form state)
        'selected_fiscal_year': fiscal_year,
        'selected_fiscal_period': fiscal_period,
        'selected_min_amount': min_amount,
        'selected_max_amount': max_amount,
        'selected_reference_number': reference_number,
        'selected_support_status': support_status,

        'chart_data': {
            'labels': ['Supported', 'Unsupported'],
            'data': [project.supported_transactions_number, project.unsupported_transactions_number],
        }
    }

    return render(request, 'reports/project_report.html', context)

def project_report_excel_export(request, project_id):
    """Export filtered report data to Excel with tiered structure"""
    project = get_object_or_404(Project, project_id=project_id)

    # Apply the same filtering logic as the main view
    supporting_docs_queryset = SupportingDocument.objects.filter(project=project)

    # Get filter parameters from request
    fiscal_year = request.GET.get('fiscal_year', '')
    fiscal_period = request.GET.get('fiscal_period', '')
    min_amount = request.GET.get('min_amount', '')
    max_amount = request.GET.get('max_amount', '')
    reference_number = request.GET.get('reference_number', '')
    support_status = request.GET.get('support_status', '')

    # Apply filters (same logic as main view)
    filtered_docs = supporting_docs_queryset

    if fiscal_year:
        filtered_docs = filtered_docs.filter(fiscal_year=fiscal_year)

    if fiscal_period:
        filtered_docs = filtered_docs.filter(fiscal_period=fiscal_period)

    if min_amount:
        try:
            min_amount_decimal = float(min_amount)
            filtered_docs = filtered_docs.filter(transaction_value__gte=min_amount_decimal)
        except ValueError:
            pass

    if max_amount:
        try:
            max_amount_decimal = float(max_amount)
            filtered_docs = filtered_docs.filter(transaction_value__lte=max_amount_decimal)
        except ValueError:
            pass

    if reference_number:
        filtered_docs = filtered_docs.filter(iddoc__icontains=reference_number.strip())

    if support_status:
        if support_status == 'supported':
            filtered_docs = filtered_docs.filter(supported=True)
        elif support_status == 'unsupported':
            filtered_docs = filtered_docs.filter(supported=False)

    filtered_docs = filtered_docs.select_related().prefetch_related('documents').order_by(
        '-fiscal_year', '-fiscal_period', 'batchnbr', 'entrynbr'
    )

    # Get comments for filtered documents
    filtered_comments = Comments.objects.filter(project=project)
    if filtered_docs.exists():
        batch_entry_combinations = [(doc.batchnbr, doc.entrynbr) for doc in filtered_docs]
        comment_filters = Q()
        for batchnbr, entrynbr in batch_entry_combinations:
            comment_filters |= Q(batchnbr=batchnbr, entrynbr=entrynbr)
        filtered_comments = filtered_comments.filter(comment_filters)

    filtered_comments = filtered_comments.select_related('user').order_by('-timestamp')

    # Get document files for filtered documents
    doc_files = SupportingDocumentFile.objects.filter(
        batch_support__in=filtered_docs
    ).select_related('batch_support').order_by('batch_support__batchnbr', 'batch_support__entrynbr')

    # Create Excel workbook
    wb = openpyxl.Workbook()

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, color="333333")
    subheader_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    # Sheet 1: Tiered Report (Main Sheet)
    ws_main = wb.active
    ws_main.title = "Tiered Report"

    # Main headers
    main_headers = [
        "Document Info", "", "", "", "Files & Comments", "", ""
    ]
    sub_headers = [
        "Batch #", "Entry #", "Reference #", "Transaction Value", "Support Status",
        "Content/File Name", "Date", "Additional Info"
    ]

    # Add main headers
    for col_idx, header in enumerate(main_headers, 1):
        cell = ws_main.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Add sub headers
    for col_idx, header in enumerate(sub_headers, 1):
        cell = ws_main.cell(row=2, column=col_idx, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Merge main header cells
    ws_main.merge_cells('A1:E1')
    ws_main.merge_cells('F1:H1')

    current_row = 3

    # Process each document with its associated files and comments
    for doc in filtered_docs:
        # Get files for this document
        doc_files_for_doc = doc_files.filter(batch_support=doc)

        # Get comments for this document
        comments_for_doc = filtered_comments.filter(batchnbr=doc.batchnbr, entrynbr=doc.entrynbr)

        # Determine how many rows we need for this document
        file_count = doc_files_for_doc.count()
        comment_count = comments_for_doc.count()
        rows_needed = max(1, file_count + comment_count)

        # Document info (spans multiple rows if needed)
        start_row = current_row

        # Add document basic info
        ws_main.cell(row=current_row, column=1, value=doc.batchnbr).border = thick_border
        ws_main.cell(row=current_row, column=2, value=doc.entrynbr).border = thick_border
        ws_main.cell(row=current_row, column=3, value=doc.iddoc.strip()).border = thick_border

        # Transaction value with formatting
        value_cell = ws_main.cell(row=current_row, column=4, value=float(doc.transaction_value))
        value_cell.number_format = 'ZMW #,##0.00'
        value_cell.border = thick_border

        # Support status with color coding
        status_cell = ws_main.cell(row=current_row, column=5, value="Supported" if doc.supported else "Unsupported")
        status_cell.border = thick_border
        if doc.supported:
            status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

        # Add files for this document
        file_row = current_row
        for doc_file in doc_files_for_doc:
            ws_main.cell(row=file_row, column=6, value=doc_file.document.name.split('/')[
                -1] if doc_file.document else doc_file.document_name).border = border
            ws_main.cell(row=file_row, column=7, value=doc_file.uploaded_at.strftime(
                "%Y-%m-%d %H:%M") if doc_file.uploaded_at else "").border = border
            ws_main.cell(row=file_row, column=8,
                         value=doc_file.document.name if doc_file.document else "").border = border
            file_row += 1

        # Add comments for this document
        comment_row = file_row
        for comment in comments_for_doc:
            # Comment text with wrapping
            comment_cell = ws_main.cell(row=comment_row, column=6, value=comment.text)
            comment_cell.border = border
            comment_cell.alignment = Alignment(wrap_text=True, vertical='top')

            ws_main.cell(row=comment_row, column=7, value=comment.timestamp.strftime("%Y-%m-%d %H:%M")).border = border

            # Author name in additional info
            author_name = comment.user.get_full_name() or comment.user.username
            ws_main.cell(row=comment_row, column=8,
                         value=f"By: {author_name} | Source: {comment.source}").border = border
            comment_row += 1

        # If we have multiple rows for this document, merge the document info cells vertically
        if rows_needed > 1:
            end_row = current_row + rows_needed - 1

            # Merge document info cells
            if start_row != end_row:
                ws_main.merge_cells(f'A{start_row}:A{end_row}')
                ws_main.merge_cells(f'B{start_row}:B{end_row}')
                ws_main.merge_cells(f'C{start_row}:C{end_row}')
                ws_main.merge_cells(f'D{start_row}:D{end_row}')
                ws_main.merge_cells(f'E{start_row}:E{end_row}')

                # Center align the merged cells
                for col in range(1, 6):
                    ws_main.cell(row=start_row, column=col).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')

        # If no files or comments, add a row indicating this
        if file_count == 0 and comment_count == 0:
            ws_main.cell(row=current_row, column=6, value="No files or comments").border = border
            ws_main.cell(row=current_row, column=7, value="").border = border
            ws_main.cell(row=current_row, column=8, value="").border = border

        current_row += max(1, rows_needed)

        # Add empty row between documents for better readability
        current_row += 1

    # Auto-adjust column widths
    column_widths = [12, 12, 20, 18, 15, 35, 18, 30]
    for col_idx, width in enumerate(column_widths, 1):
        ws_main.column_dimensions[get_column_letter(col_idx)].width = width

    # Set row heights for better readability
    ws_main.row_dimensions[1].height = 25
    ws_main.row_dimensions[2].height = 20

    # Sheet 2: Project Summary
    ws_summary = wb.create_sheet("Project Summary")

    # Project summary data
    summary_data = [
        ["Project Report Summary", ""],
        ["", ""],
        ["Project Name", project.project_name],
        ["Description", project.description or "N/A"],
        ["Created Date", project.created_at.strftime("%Y-%m-%d") if project.created_at else "N/A"],
        ["Last Updated", project.updated_at.strftime("%Y-%m-%d") if project.updated_at else "N/A"],
        ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Filters Applied", ""],
        ["Fiscal Year", fiscal_year or "All"],
        ["Fiscal Period", fiscal_period or "All"],
        ["Min Amount", f"ZMW {min_amount}" if min_amount else "No limit"],
        ["Max Amount", f"ZMW {max_amount}" if max_amount else "No limit"],
        ["Reference Number", reference_number or "All"],
        ["Support Status", support_status.title() if support_status else "All"],
        ["", ""],
        ["Summary Statistics", ""],
        ["Total Documents (Filtered)", filtered_docs.count()],
        ["Supported Documents", filtered_docs.filter(supported=True).count()],
        ["Unsupported Documents", filtered_docs.filter(supported=False).count()],
        ["Total Value (Filtered)", f"ZMW {sum(doc.transaction_value for doc in filtered_docs):.2f}"],
        ["Total Files", doc_files.count()],
        ["Total Comments", filtered_comments.count()],
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)

        if row_idx == 1:  # Title row
            ws_summary.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
        elif label and not value:  # Section headers
            ws_summary.cell(row=row_idx, column=1).font = Font(bold=True)

    # Auto-adjust column widths for summary
    for column in ws_summary.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width

    # Sheet 3: Documents Only (for quick reference)
    ws_docs = wb.create_sheet("Documents Only")

    # Document headers
    doc_headers = [
        "Batch Number", "Entry Number", "Reference Number", "Fiscal Year",
        "Fiscal Period", "Transaction Value", "Support Status",
        "Files Count", "Comments Count", "Created Date", "Updated Date"
    ]

    # Add headers
    for col_idx, header in enumerate(doc_headers, 1):
        cell = ws_docs.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Add document data
    for row_idx, doc in enumerate(filtered_docs, 2):
        ws_docs.cell(row=row_idx, column=1, value=doc.batchnbr).border = border
        ws_docs.cell(row=row_idx, column=2, value=doc.entrynbr).border = border
        ws_docs.cell(row=row_idx, column=3, value=doc.iddoc.strip()).border = border
        ws_docs.cell(row=row_idx, column=4, value=doc.fiscal_year).border = border
        ws_docs.cell(row=row_idx, column=5, value=doc.fiscal_period).border = border

        # Format transaction value
        value_cell = ws_docs.cell(row=row_idx, column=6, value=float(doc.transaction_value))
        value_cell.number_format = 'ZMW #,##0.00'
        value_cell.border = border

        ws_docs.cell(row=row_idx, column=7, value="Supported" if doc.supported else "Unsupported")
        status_cell = ws_docs.cell(row=row_idx, column=7)
        status_cell.border = border
        if doc.supported:
            status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

        # Count files and comments for this document
        files_count = doc_files.filter(batch_support=doc).count()
        comments_count = filtered_comments.filter(batchnbr=doc.batchnbr, entrynbr=doc.entrynbr).count()

        ws_docs.cell(row=row_idx, column=8, value=files_count).border = border
        ws_docs.cell(row=row_idx, column=9, value=comments_count).border = border

        ws_docs.cell(row=row_idx, column=10,
                     value=doc.created_at.strftime("%Y-%m-%d") if doc.created_at else "").border = border
        ws_docs.cell(row=row_idx, column=11,
                     value=doc.updated_at.strftime("%Y-%m-%d") if doc.updated_at else "").border = border

    # Auto-adjust column widths for documents sheet
    for column in ws_docs.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 25)
        ws_docs.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Generate filename with timestamp and filters
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filter_suffix = ""
    if any([fiscal_year, fiscal_period, min_amount, max_amount, reference_number, support_status]):
        filter_suffix = "_filtered"

    filename = f"{project.project_name.replace(' ', '_')}_tiered_report{filter_suffix}_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)
    return response
# Add this to your urls.py
from django.contrib.auth.decorators import login_required
@login_required
def project_report_pdf(request, project_id):
    """View for downloading the report as PDF using xhtml2pdf"""
    project = get_object_or_404(Project, project_id=project_id)

    # Same data preparation as in the view function
    documents = SupportingDocument.objects.filter(project=project)
    batch_entries = documents.values_list('batchnbr', 'entrynbr').distinct()

    comments_by_batch = {}
    for batchnbr, entrynbr in batch_entries:
        batch_key = f"{batchnbr}-{entrynbr}"
        # You may need to adjust this depending on how your Comments model stores batch references
        comments_by_batch[batch_key] = Comments.objects.filter(
            entrynbr=entrynbr, batchnbr=batchnbr
        ).order_by('-timestamp')

    total_transactions = project.get_total_transactions()
    support_percentage = (
        project.supported_transactions_number / total_transactions * 100
    ) if total_transactions > 0 else 0

    context = {
        'project': project,
        'documents': documents,
        'comments_by_batch': comments_by_batch,
        'support_percentage': support_percentage,
        'report_date': datetime.now(),
        'base_url': request.build_absolute_uri('/')[:-1]  # For proper asset loading in PDF
    }

    # Render the template to a string
    template = get_template('reports/project_report_pdf.html')
    html = template.render(context)

    # Create a file-like buffer to receive PDF data
    buffer = BytesIO()

    # Create the PDF object, using the BytesIO object as its "file"
    pisa_status = pisa.CreatePDF(
        html,  # the HTML to convert
        dest=buffer,  # the file-like destination
        encoding='utf-8'  # encoding of the source HTML
    )

    # If the conversion failed, return a response indicating error
    if pisa_status.err:
        return HttpResponse('PDF generation error')

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="project_report_{project.project_id}.pdf"'

    return response